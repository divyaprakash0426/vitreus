"""TDD tests for one-shot analyze --output and XLSX export."""
import json
from pathlib import Path

import pytest
from typer.testing import CliRunner

from interfaces.cli import app

runner = CliRunner()

MOCK_MANIFEST = json.dumps({
    "model": {"primary": "gemma4:31b", "drafter": "gemma4:4b", "rationale": "test"},
    "actions": [
        {"type": "highlight", "range": "Sheet1!A2:C2", "color": "#f97316", "reason": "over budget"},
        {"type": "write_value", "cell": "Sheet1!C2", "value": "OVER BUDGET"},
        {"type": "formula", "cell": "Sheet1!C4", "formula": "=SUM(B2:B3)", "reason": "total"},
    ],
})

CSV_DATA = "Name,Score,Notes\nAda,91,\nLinus,65,\nTotal,,\n"


# ─── One-shot analyze --output ────────────────────────────────────────────────

def test_analyze_with_output_applies_manifest_in_one_step(tmp_path: Path, monkeypatch):
    """analyze --output saves modified file; no separate apply-manifest needed."""
    csv_path = tmp_path / "data.csv"
    out_path = tmp_path / "data_applied.csv"
    csv_path.write_text(CSV_DATA, encoding="utf-8")
    monkeypatch.setattr("core.reasoning.GoogleAIBackend.call", lambda self, p: MOCK_MANIFEST)

    result = runner.invoke(
        app,
        ["analyze", str(csv_path), "any query", "--backend", "google", "--output", str(out_path)],
        env={"GEMINI_API_KEY": "dummy"},
    )

    assert result.exit_code == 0
    summary = json.loads(result.stdout)
    assert summary["applied"] == 3
    assert summary["saved"] == str(out_path)
    assert out_path.exists()


def test_analyze_with_output_csv_warns_about_limitations(tmp_path: Path, monkeypatch):
    """analyze --output *.csv prints a limitation warning to stderr."""
    csv_path = tmp_path / "data.csv"
    out_path = tmp_path / "data_applied.csv"
    csv_path.write_text(CSV_DATA, encoding="utf-8")
    monkeypatch.setattr("core.reasoning.GoogleAIBackend.call", lambda self, p: MOCK_MANIFEST)

    result = runner.invoke(
        app,
        ["analyze", str(csv_path), "any query", "--backend", "google", "--output", str(out_path)],
        env={"GEMINI_API_KEY": "dummy"},
    )

    assert "CSV" in result.stderr
    assert "color" in result.stderr.lower() or "highlight" in result.stderr.lower()


def test_analyze_without_output_still_prints_manifest_json(tmp_path: Path):
    """analyze without --output prints the manifest JSON (existing behaviour)."""
    csv_path = tmp_path / "data.csv"
    csv_path.write_text(CSV_DATA, encoding="utf-8")

    result = runner.invoke(app, ["analyze", str(csv_path), "Highlight rows that need review"])

    assert result.exit_code == 0
    manifest = json.loads(result.stdout)
    assert "actions" in manifest


# ─── XLSX writer unit tests ───────────────────────────────────────────────────

openpyxl = pytest.importorskip("openpyxl", reason="openpyxl not installed")


def test_workbook_snapshot_save_xlsx_writes_cell_values(tmp_path: Path):
    from core.driver import CellFormat, WorkbookSnapshot

    snapshot = WorkbookSnapshot(
        sheets={"Sheet1": [["Name", "Score", "Notes"], ["Ada", 91, "pass"], ["Linus", 65, ""]]}
    )
    out = tmp_path / "out.xlsx"
    snapshot.save_xlsx(str(out), sheet_name="Sheet1", formats={})

    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    assert ws["A1"].value == "Name"
    assert ws["B2"].value == 91
    assert ws["A3"].value == "Linus"


def test_workbook_snapshot_save_xlsx_applies_highlight_colors(tmp_path: Path):
    from core.driver import CellFormat, WorkbookSnapshot

    snapshot = WorkbookSnapshot(
        sheets={"Sheet1": [["Name", "Score"], ["Ada", 91], ["Linus", 65]]}
    )
    formats = {
        "Sheet1!A3": CellFormat(background="#f97316"),
        "Sheet1!B3": CellFormat(background="#f97316"),
    }
    out = tmp_path / "colored.xlsx"
    snapshot.save_xlsx(str(out), sheet_name="Sheet1", formats=formats)

    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    fill = ws["A3"].fill
    assert fill.fgColor.rgb.upper().endswith("F97316")


def test_workbook_snapshot_save_xlsx_writes_formula_strings(tmp_path: Path):
    from core.driver import WorkbookSnapshot

    snapshot = WorkbookSnapshot(
        sheets={"Sheet1": [["Name", "Score", "Total"], ["Ada", 91, ""], ["", "", "=SUM(B2:B2)"]]}
    )
    out = tmp_path / "formula.xlsx"
    snapshot.save_xlsx(str(out), sheet_name="Sheet1", formats={})

    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    assert ws["C3"].value == "=SUM(B2:B2)"


# ─── XLSX full pipeline via CLI ───────────────────────────────────────────────

def test_analyze_with_output_xlsx_saves_values_and_colors(tmp_path: Path, monkeypatch):
    """analyze --output *.xlsx writes cell values AND highlight colors."""
    csv_path = tmp_path / "data.csv"
    out_path = tmp_path / "data_applied.xlsx"
    csv_path.write_text(CSV_DATA, encoding="utf-8")
    monkeypatch.setattr("core.reasoning.GoogleAIBackend.call", lambda self, p: MOCK_MANIFEST)

    result = runner.invoke(
        app,
        ["analyze", str(csv_path), "any query", "--backend", "google", "--output", str(out_path)],
        env={"GEMINI_API_KEY": "dummy"},
    )

    assert result.exit_code == 0
    assert out_path.exists()
    assert result.stderr == ""  # no warning for xlsx

    wb = openpyxl.load_workbook(str(out_path))
    ws = wb.active
    assert ws["C2"].value == "OVER BUDGET"
    assert ws["C4"].value == "=SUM(B2:B3)"
    fill = ws["A2"].fill
    assert fill.fgColor.rgb.upper().endswith("F97316")


def test_workbook_snapshot_loads_xlsx_single_sheet(tmp_path: Path):
    """from_xlsx reads cell values from the first sheet by default."""
    import openpyxl as xl
    wb = xl.Workbook(); ws = wb.active; ws.title = "Data"
    ws.append(["Name", "Score"]); ws.append(["Ada", 91]); ws.append(["Linus", 65])
    path = str(tmp_path / "wb.xlsx"); wb.save(path)

    from core.driver import WorkbookSnapshot
    snapshot = WorkbookSnapshot.from_xlsx(path)
    assert snapshot.sheets["Data"][0] == ["Name", "Score"]
    assert snapshot.sheets["Data"][1] == ["Ada", 91]


def test_workbook_snapshot_loads_xlsx_all_sheets(tmp_path: Path):
    """from_xlsx with all_sheets=True loads every sheet."""
    import openpyxl as xl
    wb = xl.Workbook(); wb.active.title = "S1"; wb.create_sheet("S2")
    wb["S1"].append(["A"]); wb["S2"].append(["B"])
    path = str(tmp_path / "wb.xlsx"); wb.save(path)

    from core.driver import WorkbookSnapshot
    snapshot = WorkbookSnapshot.from_xlsx(path, all_sheets=True)
    assert "S1" in snapshot.sheets and "S2" in snapshot.sheets


def test_analyze_command_accepts_xlsx_input(tmp_path: Path):
    """analyze works when given an .xlsx file instead of .csv."""
    import openpyxl as xl
    wb = xl.Workbook(); ws = wb.active; ws.title = "Sheet1"
    ws.append(["Name", "Score"]); ws.append(["Ada", 91]); ws.append(["Linus", 65])
    xlsx_path = tmp_path / "data.xlsx"; wb.save(str(xlsx_path))

    result = CliRunner().invoke(app, ["analyze", str(xlsx_path), "Highlight rows that need review"])
    assert result.exit_code == 0
    manifest = json.loads(result.stdout)
    assert "actions" in manifest


def test_test_workbook_xlsx_has_three_sheets():
    """The generated examples/test_workbook.xlsx has Sales, Expenses, HR_Reviews sheets."""
    import openpyxl as xl
    wb = xl.load_workbook("examples/test_workbook.xlsx")
    assert set(wb.sheetnames) == {"Sales", "Expenses", "HR_Reviews"}
    assert wb["Sales"].max_row > 20
    assert wb["Expenses"].max_row > 20
    assert wb["HR_Reviews"].max_row > 15


def test_analyze_with_output_xlsx_no_separate_highlights_sidecar(tmp_path: Path, monkeypatch):
    """No _highlights.json sidecar should be created when saving to xlsx."""
    csv_path = tmp_path / "data.csv"
    out_path = tmp_path / "data_applied.xlsx"
    csv_path.write_text(CSV_DATA, encoding="utf-8")
    monkeypatch.setattr("core.reasoning.GoogleAIBackend.call", lambda self, p: MOCK_MANIFEST)

    runner.invoke(
        app,
        ["analyze", str(csv_path), "any query", "--backend", "google", "--output", str(out_path)],
        env={"GEMINI_API_KEY": "dummy"},
    )

    sidecar = tmp_path / "data_applied_highlights.json"
    assert not sidecar.exists()
